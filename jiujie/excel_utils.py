import os

from loguru import logger
from openpyxl import load_workbook
import re
from jiujie.jiujie_item_init import jiujie_item_default


class ExcelOperation:
    def __init__(self, filename, sheet_name):
        self.filename = filename
        self.ws = None
        self.sheet_name = sheet_name
        self.init_excel()
        self.jiujie_item = jiujie_item_default
        self.real_item_content = {}

    def fetch_title_value(self):
        title_content = self.jiujie_item['title']
        for key, value_loc in title_content.items():
            each_value = self.ws[value_loc].value  # 获取该坐标的值,然后进行原地更新
            title_content[key] = each_value if each_value else ''  # 获取该坐标的值,然后进行原地更新

    def calculate_next_cell(self, curr_cell):
        """计算下一行相同单元的数据"""
        cell_result = re.search('([A-Z]+)(\\d+)', curr_cell, flags=re.I)
        # if cell_result is None:
        #     raise Exception(f'单元格坐标出错,无法提取其中的行号和列号坐标为{curr_cell}')
        column = cell_result.group(1)
        row = cell_result.group(2)
        next_row = int(row) + 1
        next_cell = f'{column}{next_row}'
        return next_cell

    def guess_max_row(self, item_no_loc):
        """
        直接调用self.ws.max_row 可能得不到最大行
        :return:
        """
        count = 0
        while True:
            item_value = self.ws[item_no_loc].value  # 获取单元格数值
            if not item_value:  # 如果不存在该数据
                break
            # 计算下一个单元格
            item_no_loc = self.calculate_next_cell(item_no_loc)
            count += 1
        return count

    def handle_all_item_data(self, item_data):
        zip_item = zip(*item_data)
        item_list = list(map(list, zip_item))  # 将每个元素变成list 类型
        logger.info(item_list)
        count = 0
        for key in self.real_item_content.keys():
            self.real_item_content[key] = item_list[count]
            count += 1
        logger.info(self.real_item_content)  # 因为字典是无序的, 未来数据可能会错位
        self.jiujie_item['item'].update(self.real_item_content)

    def fetch_all_item(self, item_no_loc, row_item):
        item_start_result = re.search('(\\d+)', item_no_loc, flags=re.I)
        item_start = int(item_start_result.group(1))
        item_stop = item_start + row_item
        all_row_data = []
        for row_num in range(item_start, item_stop):
            row_content = self.ws[row_num]  # 最大列遇到None 才停止
            row_data = []
            logger.info(f'表体的单元格为:{row_content}')
            for each_cell in row_content:
                each_cell_value = each_cell.value
                if each_cell_value is not None:
                    row_data.append(each_cell_value)
            # row_data 为一行数据
            all_row_data.append(row_data)
        logger.info(f'提取到的{row_item}行数据: {all_row_data}')
        # logger.info(f'提取到的{row_item}行数据: {len(all_row_data[0])}')
        return all_row_data

    def fetch_item_value(self):
        item_content = self.jiujie_item['item']  # 用来获取坐标
        self.real_item_content = {key: [] for key in item_content.keys()}  # 用来赋值
        item_no_loc = item_content['item_no']
        row_item = self.guess_max_row(item_no_loc)  # item有几行数据
        logger.info(f'表体的行数为: {row_item}')
        # 因为获取到的数据和real_item_content 的位置一一对应
        item_data = self.fetch_all_item(item_no_loc, row_item)
        self.handle_all_item_data(item_data)

    def init_excel(self, ):
        """打开excel,获取报关单子表的sheet 对象,并将其赋值给self.ws"""
        wb = load_workbook(self.filename, read_only=True)  # 为脱离页眉页脚的警告,加入read_only参数
        self.ws = wb[self.sheet_name]  # 加载报关单子表

    def main(self):
        self.fetch_title_value()
        self.fetch_item_value()
        pass


if __name__ == '__main__':
    # file_path = os.path.join(TEMPLATE_DIR, 'CN_20201117-142628787.xlsx')
    file_path = os.path.join('单一窗口平台货物申报导入模板1- 数据测试.xlsx')
    sheet_name = '基本信息&商品信息'
    eo = ExcelOperation(file_path, sheet_name)
    eo.main()
    # eo.fetch_cell_value()
