import os

from loguru import logger
from openpyxl import load_workbook

from hongming_backend.constants import TEMPLATE_DIR


class ExcelOperation:
    def __init__(self, filename, sheet_name='Sheet1'):
        self.filename = filename
        self.ws = None
        self.sheet_name = sheet_name
        self.init_excel()

    def init_excel(self, ):
        """打开excel,获取报关单子表的sheet 对象,并将其赋值给self.ws"""
        wb = load_workbook(self.filename, read_only=True)  # 为脱离页眉页脚的警告,加入read_only参数
        self.ws = wb[self.sheet_name]  # 加载报关单子表

    def judge_key(self, cell_1):
        """用来判断当前单元格是否为大键,例如title,type"""

        pass

    @logger.catch
    def fetch_cell_value(self):
        """
        从A1 ,B1 开始读取,返回键值对数据,返回字典数据,第一列为键,第二列为值
        """
        max_row = self.ws.max_row
        final_data = {}
        for row in range(1, max_row + 1):
            key = self.ws.cell(row, 1).value
            value = self.ws.cell(row, 2).value
            final_data[key] = value if value else ""
        print(final_data)
        return final_data


if __name__ == '__main__':
    # file_path = os.path.join(TEMPLATE_DIR, 'CN_20201117-142628787.xlsx')
    file_path = os.path.join(TEMPLATE_DIR, '单一窗口平台货物申报导入模板1- 数据测试.xlsx')
    eo = ExcelOperation(file_path)
    eo.fetch_cell_value()
